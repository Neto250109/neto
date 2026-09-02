"""
ETL — Etapa 1: Importação
Lê os três arquivos oficiais de divulgação do IDEB 2025 (INEP/MEC), cujo
cabeçalho ocupa 4 linhas mescladas (título, rótulo do bloco, sub-rótulo,
nome técnico da variável) antes do início dos dados na linha 6.

A linha 5 (índice 4) contém os nomes técnicos padronizados pelo INEP
(ex.: VL_OBSERVADO_2025, VL_APROVACAO_2023_SI_4) — usamos esses nomes como
chave programática confiável para reconstruir a estrutura, em vez de tentar
inferir os nomes a partir dos rótulos mesclados (que se repetem e têm
células vazias por mesclagem).

Saída: um arquivo parquet "long" por etapa em dados/processado/, com uma
linha por (UF, Município, Rede, Etapa, Ano, Indicador), preservando o
valor bruto original antes de qualquer conversão.
"""
import re
import json
import openpyxl
import pandas as pd
from pathlib import Path

RAW_DIR = Path(__file__).resolve().parents[1] / "dados" / "bruto"
OUT_DIR = Path(__file__).resolve().parents[1] / "dados" / "processado"
AUDIT_DIR = Path(__file__).resolve().parents[1] / "audit"
OUT_DIR.mkdir(parents=True, exist_ok=True)
AUDIT_DIR.mkdir(parents=True, exist_ok=True)

FILES = {
    "Anos Iniciais": RAW_DIR / "divulgacao_anos_iniciais_municipios_2025.xlsx",
    "Anos Finais": RAW_DIR / "divulgacao_anos_finais_municipios_2025.xlsx",
    "Ensino Médio": RAW_DIR / "divulgacao_ensino_medio_municipios_2025.xlsx",
}

ID_COLS = {"SG_UF": "uf", "CO_MUNICIPIO": "co_municipio", "NO_MUNICIPIO": "nome_municipio", "REDE": "rede"}

# Mapeia o nome técnico (linha 5 do cabeçalho) para (grupo do indicador, nome do indicador)
VAR_RE = re.compile(r"^VL_(APROVACAO|INDICADOR_REND|NOTA_MATEMATICA|NOTA_PORTUGUES|NOTA_MEDIA|OBSERVADO|PROJECAO)_(\d{4})(_SI_4)?$")

INDICADOR_LABELS = {
    "APROVACAO": "Taxa de Aprovação",
    "INDICADOR_REND": "Indicador de Rendimento",
    "NOTA_MATEMATICA": "Nota SAEB - Matemática",
    "NOTA_PORTUGUES": "Nota SAEB - Língua Portuguesa",
    "NOTA_MEDIA": "Nota Média Padronizada",
    "OBSERVADO": "IDEB",
    "PROJECAO": "Meta IDEB",
}


def read_header(ws):
    """Lê as 5 primeiras linhas e retorna a lista de nomes técnicos (linha 5)."""
    rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    header_tecnico = rows[4]
    return header_tecnico


def extract_sheet(fn: Path, etapa: str):
    wb = openpyxl.load_workbook(fn, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_tecnico = read_header(ws)

    # localizar colunas de identificação e colunas de indicador válidas
    id_idx = {}
    value_cols = []  # list of (col_index, indicador_grupo, ano, indicador_label)
    for i, name in enumerate(header_tecnico):
        if name is None:
            continue
        name = str(name).strip()
        if name in ID_COLS:
            id_idx[ID_COLS[name]] = i
            continue
        m = VAR_RE.match(name)
        if not m:
            continue  # colunas de detalhe por série/ano individual (ex.: VL_APROVACAO_2023_1) são ignoradas
        grupo, ano, is_total = m.group(1), int(m.group(2)), m.group(3)
        if grupo == "APROVACAO" and not is_total:
            continue  # mantém apenas o total da etapa (sufixo _SI_4), não o detalhe por série/ano escolar
        value_cols.append((i, grupo, ano, INDICADOR_LABELS[grupo]))

    assert set(["uf", "co_municipio", "nome_municipio", "rede"]).issubset(id_idx), f"colunas de identificação não encontradas em {fn}"

    records = []
    n_rows = 0
    for row in ws.iter_rows(min_row=6, values_only=True):
        if row[id_idx["co_municipio"]] is None:
            continue
        n_rows += 1
        uf = row[id_idx["uf"]]
        co_mun = row[id_idx["co_municipio"]]
        nome_mun = row[id_idx["nome_municipio"]]
        rede = row[id_idx["rede"]]
        for col_i, grupo, ano, label in value_cols:
            raw = row[col_i]
            records.append((uf, co_mun, nome_mun, rede, etapa, ano, label, grupo, raw))
    wb.close()
    df = pd.DataFrame(records, columns=["uf", "co_municipio", "nome_municipio", "rede", "etapa", "ano", "indicador", "indicador_grupo", "valor_bruto"])
    # valor_bruto mistura float/int/str (INEP grava número, "-" ou "ND"); mantemos como texto
    # bruto fiel ao original para rastreabilidade, e convertemos na etapa de limpeza (02).
    df["valor_bruto"] = df["valor_bruto"].apply(lambda v: v if v is None else str(v))
    return df, n_rows


def main():
    catalogo = {}
    for etapa, fn in FILES.items():
        df, n_rows = extract_sheet(fn, etapa)
        out_fn = OUT_DIR / f"long_{etapa.lower().replace(' ', '_').replace('í','i').replace('é','e')}.parquet"
        # normaliza nome de arquivo (sem acentos) manualmente
        slug = {"Anos Iniciais": "anos_iniciais", "Anos Finais": "anos_finais", "Ensino Médio": "ensino_medio"}[etapa]
        out_fn = OUT_DIR / f"long_{slug}.parquet"
        df.to_parquet(out_fn, index=False)
        catalogo[etapa] = {
            "arquivo_origem": fn.name,
            "linhas_dados_lidas": n_rows,
            "registros_long": len(df),
            "municipios_distintos": int(df["co_municipio"].nunique()),
            "redes_distintas": sorted(df["rede"].dropna().unique().tolist()),
            "anos_distintos": sorted(int(a) for a in df["ano"].dropna().unique().tolist()),
            "indicadores_distintos": sorted(df["indicador"].dropna().unique().tolist()),
            "arquivo_saida": str(out_fn.relative_to(OUT_DIR.parents[0])),
        }
        print(f"[{etapa}] linhas lidas={n_rows} registros_long={len(df)} municipios={catalogo[etapa]['municipios_distintos']}")

    with open(AUDIT_DIR / "catalogo_fontes.json", "w", encoding="utf-8") as f:
        json.dump(catalogo, f, ensure_ascii=False, indent=2)
    print("Catálogo salvo em audit/catalogo_fontes.json")


if __name__ == "__main__":
    main()
